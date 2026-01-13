"""
Retail Store Startup Excel Bundle Generator
生成带公式和专业样式的 Excel 模板
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment,
    NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
import os

# ============================================
# 样式定义
# ============================================

# 颜色定义
COLORS = {
    'primary': '1E3A5F',      # 深蓝色 - 主色
    'secondary': '2E86AB',     # 蓝色
    'accent': 'F5A623',        # 橙色 - 强调色
    'success': '27AE60',       # 绿色
    'warning': 'F39C12',       # 黄色
    'danger': 'E74C3C',        # 红色
    'light': 'F8F9FA',         # 浅灰
    'white': 'FFFFFF',
    'text': '2C3E50',
    'input_bg': 'FFFDE7',      # 浅黄色 - 输入区域背景
}

# 字体
FONTS = {
    'title': Font(name='Arial', size=18, bold=True, color=COLORS['white']),
    'section': Font(name='Arial', size=14, bold=True, color=COLORS['white']),
    'header': Font(name='Arial', size=11, bold=True, color=COLORS['text']),
    'normal': Font(name='Arial', size=10, color=COLORS['text']),
    'input': Font(name='Arial', size=10, color=COLORS['text']),
    'total': Font(name='Arial', size=11, bold=True, color=COLORS['primary']),
    'currency': Font(name='Arial', size=10, color=COLORS['text']),
    'link': Font(name='Arial', size=10, color=COLORS['secondary'], underline='single'),
}

# 填充色
FILLS = {
    'primary': PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid'),
    'secondary': PatternFill(start_color=COLORS['secondary'], end_color=COLORS['secondary'], fill_type='solid'),
    'header': PatternFill(start_color=COLORS['light'], end_color=COLORS['light'], fill_type='solid'),
    'input': PatternFill(start_color=COLORS['input_bg'], end_color=COLORS['input_bg'], fill_type='solid'),
    'subtotal': PatternFill(start_color='E8F4FD', end_color='E8F4FD', fill_type='solid'),
    'total': PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid'),
    'white': PatternFill(start_color=COLORS['white'], end_color=COLORS['white'], fill_type='solid'),
}

# 边框
BORDERS = {
    'thin': Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    ),
    'medium': Border(
        left=Side(style='medium', color=COLORS['primary']),
        right=Side(style='medium', color=COLORS['primary']),
        top=Side(style='medium', color=COLORS['primary']),
        bottom=Side(style='medium', color=COLORS['primary'])
    ),
}

# 对齐
ALIGNMENTS = {
    'center': Alignment(horizontal='center', vertical='center'),
    'left': Alignment(horizontal='left', vertical='center'),
    'right': Alignment(horizontal='right', vertical='center'),
    'wrap': Alignment(horizontal='left', vertical='center', wrap_text=True),
}

def set_column_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

def create_title_row(ws, row, title, cols=5):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = title
    cell.font = FONTS['title']
    cell.fill = FILLS['primary']
    cell.alignment = ALIGNMENTS['center']
    ws.row_dimensions[row].height = 35

def create_section_header(ws, row, title, cols=5):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = title
    cell.font = FONTS['section']
    cell.fill = FILLS['secondary']
    cell.alignment = ALIGNMENTS['center']
    ws.row_dimensions[row].height = 28

def create_header_row(ws, row, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = FONTS['header']
        cell.fill = FILLS['header']
        cell.border = BORDERS['thin']
        cell.alignment = ALIGNMENTS['center']
    ws.row_dimensions[row].height = 22


def create_startup_cost_calculator(output_dir):
    """创建开店成本计算器 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Startup Cost Calculator"
    
    set_column_widths(ws, {'A': 35, 'B': 40, 'C': 18, 'D': 45})
    
    create_title_row(ws, 1, "RETAIL STORE STARTUP COST CALCULATOR", 4)
    ws.cell(row=2, column=1).value = "Version 1.0 | Goodok Shopfitting | www.goodokshop.com"
    ws.cell(row=2, column=1).font = Font(size=10, italic=True, color='666666')
    ws.merge_cells('A2:D2')
    
    ws.cell(row=4, column=1).value = "Instructions: Fill in the YELLOW cells with your estimates. Formulas will calculate totals automatically."
    ws.cell(row=4, column=1).font = Font(size=10, bold=True, color=COLORS['secondary'])
    ws.merge_cells('A4:D4')
    
    row = 6
    subtotal_cells = []
    
    # Section 1: Location & Lease
    create_section_header(ws, row, "SECTION 1: LOCATION & LEASE COSTS", 4)
    row += 1
    create_header_row(ws, row, ["Category", "Item", "Estimated Cost ($)", "Notes"])
    row += 1
    
    lease_data = [
        ("Lease", "Security Deposit (First & Last Month)", 5000, "Typically 2-3 months rent"),
        ("Lease", "First Month's Rent", 2500, ""),
        ("Lease", "Broker/Agent Fee", 1250, "Usually 0.5-1 month rent"),
        ("Lease", "Lease Negotiation/Legal Fees", 500, ""),
    ]
    start_row = row
    for cat, item, cost, notes in lease_data:
        ws.cell(row=row, column=1, value=cat).border = BORDERS['thin']
        ws.cell(row=row, column=2, value=item).border = BORDERS['thin']
        cost_cell = ws.cell(row=row, column=3, value=cost)
        cost_cell.fill = FILLS['input']
        cost_cell.border = BORDERS['thin']
        cost_cell.number_format = '"$"#,##0'
        ws.cell(row=row, column=4, value=notes).border = BORDERS['thin']
        row += 1
    
    ws.cell(row=row, column=1, value="SUBTOTAL: LOCATION & LEASE")
    ws.cell(row=row, column=1).font = FONTS['total']
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    subtotal_cell = ws.cell(row=row, column=3)
    subtotal_cell.value = f"=SUM(C{start_row}:C{row-1})"
    subtotal_cell.font = FONTS['total']
    subtotal_cell.fill = FILLS['subtotal']
    subtotal_cell.border = BORDERS['thin']
    subtotal_cell.number_format = '"$"#,##0'
    subtotal_cells.append(f"C{row}")
    row += 2
    
    # Section 2: Renovation
    create_section_header(ws, row, "SECTION 2: STORE BUILD-OUT & RENOVATION", 4)
    row += 1
    create_header_row(ws, row, ["Category", "Item", "Estimated Cost ($)", "Notes"])
    row += 1
    
    reno_data = [
        ("Renovation", "Flooring", 3000, "Vinyl, tile, or hardwood"),
        ("Renovation", "Painting", 1000, ""),
        ("Renovation", "Lighting Installation", 2500, "LED fixtures recommended"),
        ("Renovation", "Electrical Work", 1500, ""),
        ("Renovation", "Plumbing (if needed)", 0, ""),
        ("Renovation", "HVAC/Climate Control", 1000, ""),
        ("Renovation", "Signage (Exterior)", 1500, ""),
        ("Renovation", "Signage (Interior)", 500, ""),
        ("Renovation", "Permits & Inspections", 300, ""),
    ]
    start_row = row
    for cat, item, cost, notes in reno_data:
        ws.cell(row=row, column=1, value=cat).border = BORDERS['thin']
        ws.cell(row=row, column=2, value=item).border = BORDERS['thin']
        cost_cell = ws.cell(row=row, column=3, value=cost)
        cost_cell.fill = FILLS['input']
        cost_cell.border = BORDERS['thin']
        cost_cell.number_format = '"$"#,##0'
        ws.cell(row=row, column=4, value=notes).border = BORDERS['thin']
        row += 1
    
    ws.cell(row=row, column=1, value="SUBTOTAL: RENOVATION")
    ws.cell(row=row, column=1).font = FONTS['total']
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    subtotal_cell = ws.cell(row=row, column=3)
    subtotal_cell.value = f"=SUM(C{start_row}:C{row-1})"
    subtotal_cell.font = FONTS['total']
    subtotal_cell.fill = FILLS['subtotal']
    subtotal_cell.border = BORDERS['thin']
    subtotal_cell.number_format = '"$"#,##0'
    subtotal_cells.append(f"C{row}")
    row += 2
    
    # Section 3: Fixtures
    create_section_header(ws, row, "SECTION 3: FIXTURES & EQUIPMENT", 4)
    row += 1
    create_header_row(ws, row, ["Item", "Qty", "Unit Price ($)", "Total ($)"])
    row += 1
    
    fixture_data = [
        ("Gondola Shelving Units (Double-sided, 4ft)", 10, 350),
        ("Wall Shelving Units (Single-sided)", 8, 250),
        ("Display Cases (Glass)", 4, 400),
        ("Checkout Counter", 1, 800),
        ("Clothing Racks (if applicable)", 0, 150),
        ("Slatwall Panels", 6, 100),
        ("Shopping Baskets", 20, 15),
        ("Shopping Carts", 5, 120),
        ("Price Tag Gun", 2, 25),
        ("Security Tags & Detacher", 1, 300),
        ("Mirrors", 4, 75),
    ]
    start_row = row
    for item, qty, price in fixture_data:
        ws.cell(row=row, column=1, value=item).border = BORDERS['thin']
        qty_cell = ws.cell(row=row, column=2, value=qty)
        qty_cell.fill = FILLS['input']
        qty_cell.border = BORDERS['thin']
        qty_cell.alignment = ALIGNMENTS['center']
        price_cell = ws.cell(row=row, column=3, value=price)
        price_cell.fill = FILLS['input']
        price_cell.border = BORDERS['thin']
        price_cell.number_format = '"$"#,##0'
        total_cell = ws.cell(row=row, column=4)
        total_cell.value = f"=B{row}*C{row}"
        total_cell.border = BORDERS['thin']
        total_cell.number_format = '"$"#,##0'
        row += 1
    
    ws.cell(row=row, column=1, value="SUBTOTAL: FIXTURES & EQUIPMENT")
    ws.cell(row=row, column=1).font = FONTS['total']
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    subtotal_cell = ws.cell(row=row, column=4)
    subtotal_cell.value = f"=SUM(D{start_row}:D{row-1})"
    subtotal_cell.font = FONTS['total']
    subtotal_cell.fill = FILLS['subtotal']
    subtotal_cell.border = BORDERS['thin']
    subtotal_cell.number_format = '"$"#,##0'
    subtotal_cells.append(f"D{row}")
    row += 2
    
    # Section 4: Technology
    create_section_header(ws, row, "SECTION 4: TECHNOLOGY & POS SYSTEM", 4)
    row += 1
    create_header_row(ws, row, ["Category", "Item", "Estimated Cost ($)", "Notes"])
    row += 1
    
    tech_data = [
        ("POS", "POS System (Hardware)", 1200, "Register, scanner, printer"),
        ("POS", "POS Software (Annual)", 600, "Square, Shopify, Clover"),
        ("Technology", "Security Camera System", 800, "4-8 cameras + DVR"),
        ("Technology", "Alarm System", 500, "Installation + first year"),
        ("Technology", "Wi-Fi Router & Setup", 150, ""),
        ("Technology", "Computer/Tablet", 500, ""),
    ]
    start_row = row
    for cat, item, cost, notes in tech_data:
        ws.cell(row=row, column=1, value=cat).border = BORDERS['thin']
        ws.cell(row=row, column=2, value=item).border = BORDERS['thin']
        cost_cell = ws.cell(row=row, column=3, value=cost)
        cost_cell.fill = FILLS['input']
        cost_cell.border = BORDERS['thin']
        cost_cell.number_format = '"$"#,##0'
        ws.cell(row=row, column=4, value=notes).border = BORDERS['thin']
        row += 1
    
    ws.cell(row=row, column=1, value="SUBTOTAL: TECHNOLOGY")
    ws.cell(row=row, column=1).font = FONTS['total']
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    subtotal_cell = ws.cell(row=row, column=3)
    subtotal_cell.value = f"=SUM(C{start_row}:C{row-1})"
    subtotal_cell.font = FONTS['total']
    subtotal_cell.fill = FILLS['subtotal']
    subtotal_cell.border = BORDERS['thin']
    subtotal_cell.number_format = '"$"#,##0'
    subtotal_cells.append(f"C{row}")
    row += 2
    
    # Section 5: Inventory
    create_section_header(ws, row, "SECTION 5: INITIAL INVENTORY", 4)
    row += 1
    create_header_row(ws, row, ["Category", "Item", "Estimated Cost ($)", "Notes"])
    row += 1
    
    inv_data = [
        ("Inventory", "Opening Stock - Category A", 10000, "Your main product category"),
        ("Inventory", "Opening Stock - Category B", 5000, "Secondary category"),
        ("Inventory", "Opening Stock - Category C", 3000, "Tertiary category"),
        ("Inventory", "Packaging Supplies", 500, "Bags, tissue, boxes"),
    ]
    start_row = row
    for cat, item, cost, notes in inv_data:
        ws.cell(row=row, column=1, value=cat).border = BORDERS['thin']
        ws.cell(row=row, column=2, value=item).border = BORDERS['thin']
        cost_cell = ws.cell(row=row, column=3, value=cost)
        cost_cell.fill = FILLS['input']
        cost_cell.border = BORDERS['thin']
        cost_cell.number_format = '"$"#,##0'
        ws.cell(row=row, column=4, value=notes).border = BORDERS['thin']
        row += 1
    
    ws.cell(row=row, column=1, value="SUBTOTAL: INVENTORY")
    ws.cell(row=row, column=1).font = FONTS['total']
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    subtotal_cell = ws.cell(row=row, column=3)
    subtotal_cell.value = f"=SUM(C{start_row}:C{row-1})"
    subtotal_cell.font = FONTS['total']
    subtotal_cell.fill = FILLS['subtotal']
    subtotal_cell.border = BORDERS['thin']
    subtotal_cell.number_format = '"$"#,##0'
    subtotal_cells.append(f"C{row}")
    row += 2
    
    # Section 6: Legal
    create_section_header(ws, row, "SECTION 6: LICENSES & LEGAL", 4)
    row += 1
    create_header_row(ws, row, ["Category", "Item", "Estimated Cost ($)", "Notes"])
    row += 1
    
    legal_data = [
        ("Legal", "Business Registration", 150, ""),
        ("Legal", "Sales Tax Permit", 0, "Often free"),
        ("Legal", "Health Permit (if food)", 300, ""),
        ("Legal", "Fire Safety Inspection", 100, ""),
        ("Legal", "Business Insurance (Annual)", 1500, "Liability + property"),
        ("Legal", "Trademark Registration", 350, "Optional but recommended"),
        ("Legal", "Attorney Consultation", 500, "Contract review"),
    ]
    start_row = row
    for cat, item, cost, notes in legal_data:
        ws.cell(row=row, column=1, value=cat).border = BORDERS['thin']
        ws.cell(row=row, column=2, value=item).border = BORDERS['thin']
        cost_cell = ws.cell(row=row, column=3, value=cost)
        cost_cell.fill = FILLS['input']
        cost_cell.border = BORDERS['thin']
        cost_cell.number_format = '"$"#,##0'
        ws.cell(row=row, column=4, value=notes).border = BORDERS['thin']
        row += 1
    
    ws.cell(row=row, column=1, value="SUBTOTAL: LICENSES & LEGAL")
    ws.cell(row=row, column=1).font = FONTS['total']
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    subtotal_cell = ws.cell(row=row, column=3)
    subtotal_cell.value = f"=SUM(C{start_row}:C{row-1})"
    subtotal_cell.font = FONTS['total']
    subtotal_cell.fill = FILLS['subtotal']
    subtotal_cell.border = BORDERS['thin']
    subtotal_cell.number_format = '"$"#,##0'
    subtotal_cells.append(f"C{row}")
    row += 2
    
    # Section 7: Marketing
    create_section_header(ws, row, "SECTION 7: MARKETING & GRAND OPENING", 4)
    row += 1
    create_header_row(ws, row, ["Category", "Item", "Estimated Cost ($)", "Notes"])
    row += 1
    
    mkt_data = [
        ("Marketing", "Logo Design", 300, ""),
        ("Marketing", "Business Cards & Flyers", 200, ""),
        ("Marketing", "Website Development", 1500, "Or use Shopify/Wix"),
        ("Marketing", "Social Media Setup", 0, "DIY"),
        ("Marketing", "Grand Opening Event", 1000, "Promotions, refreshments"),
        ("Marketing", "Local Advertising", 500, "First month"),
        ("Marketing", "Google/Meta Ads", 500, "First month"),
    ]
    start_row = row
    for cat, item, cost, notes in mkt_data:
        ws.cell(row=row, column=1, value=cat).border = BORDERS['thin']
        ws.cell(row=row, column=2, value=item).border = BORDERS['thin']
        cost_cell = ws.cell(row=row, column=3, value=cost)
        cost_cell.fill = FILLS['input']
        cost_cell.border = BORDERS['thin']
        cost_cell.number_format = '"$"#,##0'
        ws.cell(row=row, column=4, value=notes).border = BORDERS['thin']
        row += 1
    
    ws.cell(row=row, column=1, value="SUBTOTAL: MARKETING")
    ws.cell(row=row, column=1).font = FONTS['total']
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    subtotal_cell = ws.cell(row=row, column=3)
    subtotal_cell.value = f"=SUM(C{start_row}:C{row-1})"
    subtotal_cell.font = FONTS['total']
    subtotal_cell.fill = FILLS['subtotal']
    subtotal_cell.border = BORDERS['thin']
    subtotal_cell.number_format = '"$"#,##0'
    subtotal_cells.append(f"C{row}")
    row += 2
    
    # Section 8: Operating Capital
    create_section_header(ws, row, "SECTION 8: OPERATING CAPITAL (3 Months Reserve)", 4)
    row += 1
    create_header_row(ws, row, ["Item", "Monthly Cost ($)", "Months", "Total ($)"])
    row += 1
    
    opex_data = [
        ("Rent", 2500, 3),
        ("Utilities", 400, 3),
        ("Payroll", 4000, 3),
        ("Inventory Replenishment", 2000, 3),
        ("Marketing/Advertising", 300, 3),
        ("Miscellaneous", 500, 3),
    ]
    start_row = row
    for item, monthly, months in opex_data:
        ws.cell(row=row, column=1, value=item).border = BORDERS['thin']
        monthly_cell = ws.cell(row=row, column=2, value=monthly)
        monthly_cell.fill = FILLS['input']
        monthly_cell.border = BORDERS['thin']
        monthly_cell.number_format = '"$"#,##0'
        months_cell = ws.cell(row=row, column=3, value=months)
        months_cell.fill = FILLS['input']
        months_cell.border = BORDERS['thin']
        months_cell.alignment = ALIGNMENTS['center']
        total_cell = ws.cell(row=row, column=4)
        total_cell.value = f"=B{row}*C{row}"
        total_cell.border = BORDERS['thin']
        total_cell.number_format = '"$"#,##0'
        row += 1
    
    ws.cell(row=row, column=1, value="SUBTOTAL: OPERATING CAPITAL")
    ws.cell(row=row, column=1).font = FONTS['total']
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    subtotal_cell = ws.cell(row=row, column=4)
    subtotal_cell.value = f"=SUM(D{start_row}:D{row-1})"
    subtotal_cell.font = FONTS['total']
    subtotal_cell.fill = FILLS['subtotal']
    subtotal_cell.border = BORDERS['thin']
    subtotal_cell.number_format = '"$"#,##0'
    subtotal_cells.append(f"D{row}")
    row += 3
    
    # Grand Total
    create_section_header(ws, row, "GRAND TOTAL SUMMARY", 4)
    row += 1
    
    summary_items = [
        ("Location & Lease", subtotal_cells[0]),
        ("Renovation & Build-out", subtotal_cells[1]),
        ("Fixtures & Equipment", subtotal_cells[2]),
        ("Technology & POS", subtotal_cells[3]),
        ("Initial Inventory", subtotal_cells[4]),
        ("Licenses & Legal", subtotal_cells[5]),
        ("Marketing & Launch", subtotal_cells[6]),
        ("Operating Capital (3 months)", subtotal_cells[7]),
    ]
    
    for item, ref in summary_items:
        ws.cell(row=row, column=1, value=item).border = BORDERS['thin']
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ref_cell = ws.cell(row=row, column=4)
        ref_cell.value = f"={ref}"
        ref_cell.border = BORDERS['thin']
        ref_cell.number_format = '"$"#,##0'
        row += 1
    
    ws.cell(row=row, column=1, value="TOTAL STARTUP COST")
    ws.cell(row=row, column=1).font = Font(name='Arial', size=14, bold=True, color=COLORS['white'])
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1).fill = FILLS['primary']
    total_cell = ws.cell(row=row, column=4)
    total_cell.value = f"=SUM(D{row-8}:D{row-1})"
    total_cell.font = Font(name='Arial', size=14, bold=True, color=COLORS['white'])
    total_cell.fill = FILLS['primary']
    total_cell.number_format = '"$"#,##0'
    ws.row_dimensions[row].height = 30
    row += 3
    
    # Footer
    ws.cell(row=row, column=1, value="Need quality fixtures at competitive prices?")
    ws.cell(row=row, column=1).font = Font(size=12, bold=True, color=COLORS['secondary'])
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    ws.cell(row=row, column=1, value="Contact Goodok Shopfitting: www.goodokshop.com | info@goodokshop.com")
    ws.cell(row=row, column=1).font = FONTS['link']
    ws.merge_cells(f'A{row}:D{row}')
    
    ws.freeze_panes = 'A2'
    
    filepath = os.path.join(output_dir, "Retail_Startup_Cost_Calculator.xlsx")
    wb.save(filepath)
    return filepath


if __name__ == "__main__":
    output_dir = r"E:\retail-miniapp-backend\goodok-website\digital-products"
    
    print("Generating Retail Store Startup Excel Bundle...")
    
    print("  Creating Startup Cost Calculator...")
    filepath = create_startup_cost_calculator(output_dir)
    print(f"  Saved: {filepath}")
    
    print("\nExcel template generated successfully!")
