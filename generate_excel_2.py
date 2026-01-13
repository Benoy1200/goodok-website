"""
生成剩余的 Excel 模板
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import FormulaRule, CellIsRule
import os

# 样式定义
COLORS = {
    'primary': '1E3A5F', 'secondary': '2E86AB', 'accent': 'F5A623',
    'success': '27AE60', 'warning': 'F39C12', 'danger': 'E74C3C',
    'light': 'F8F9FA', 'white': 'FFFFFF', 'text': '2C3E50', 'input_bg': 'FFFDE7',
}

FONTS = {
    'title': Font(name='Arial', size=18, bold=True, color=COLORS['white']),
    'section': Font(name='Arial', size=14, bold=True, color=COLORS['white']),
    'header': Font(name='Arial', size=11, bold=True, color=COLORS['text']),
    'normal': Font(name='Arial', size=10, color=COLORS['text']),
    'total': Font(name='Arial', size=11, bold=True, color=COLORS['primary']),
    'link': Font(name='Arial', size=10, color=COLORS['secondary'], underline='single'),
}

FILLS = {
    'primary': PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid'),
    'secondary': PatternFill(start_color=COLORS['secondary'], end_color=COLORS['secondary'], fill_type='solid'),
    'header': PatternFill(start_color=COLORS['light'], end_color=COLORS['light'], fill_type='solid'),
    'input': PatternFill(start_color=COLORS['input_bg'], end_color=COLORS['input_bg'], fill_type='solid'),
    'subtotal': PatternFill(start_color='E8F4FD', end_color='E8F4FD', fill_type='solid'),
    'success': PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid'),
    'warning': PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid'),
    'danger': PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid'),
}

BORDERS = {
    'thin': Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    ),
}

ALIGNMENTS = {
    'center': Alignment(horizontal='center', vertical='center'),
    'left': Alignment(horizontal='left', vertical='center'),
}

def set_column_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

def create_title_row(ws, row, title, cols=5):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = title
    cell.font = FONTS['title']
    cell.fill = FILLS['primary']
    cell.alignment = ALIGNMENTS['center']
    ws.row_dimensions[row].height = 35

def create_section_header(ws, row, title, cols=5):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = title
    cell.font = FONTS['section']
    cell.fill = FILLS['secondary']
    cell.alignment = ALIGNMENTS['center']
    ws.row_dimensions[row].height = 28

def create_header_row(ws, row, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = FONTS['header']
        cell.fill = FILLS['header']
        cell.border = BORDERS['thin']
        cell.alignment = ALIGNMENTS['center']
    ws.row_dimensions[row].height = 22


def create_store_opening_checklist(output_dir):
    """创建开店任务清单 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Store Opening Checklist"
    
    set_column_widths(ws, {'A': 12, 'B': 50, 'C': 15, 'D': 15, 'E': 15, 'F': 35})
    
    create_title_row(ws, 1, "RETAIL STORE OPENING CHECKLIST", 6)
    ws.cell(row=2, column=1).value = "Version 1.0 | Goodok Shopfitting | www.goodokshop.com"
    ws.cell(row=2, column=1).font = Font(size=10, italic=True, color='666666')
    ws.merge_cells('A2:F2')
    
    ws.cell(row=4, column=1).value = "Instructions: Mark status as DONE when complete. Target dates are editable."
    ws.cell(row=4, column=1).font = Font(size=10, bold=True, color=COLORS['secondary'])
    ws.merge_cells('A4:F4')
    
    phases = [
        ("PHASE 1: PLANNING & RESEARCH (Weeks 1-4)", [
            ("Define your target market and niche", "Owner", "Week 1"),
            ("Research competitors in your area", "Owner", "Week 1"),
            ("Create a detailed business plan", "Owner", "Week 2"),
            ("Determine startup budget", "Owner", "Week 2"),
            ("Apply for business financing (if needed)", "Owner", "Week 2-3"),
            ("Select business structure (LLC, Corp, etc.)", "Owner", "Week 3"),
            ("Choose a business name", "Owner", "Week 3"),
            ("Register your business", "Owner", "Week 4"),
        ]),
        ("PHASE 2: LOCATION & LEASE (Weeks 4-6)", [
            ("Identify potential locations", "Owner", "Week 4"),
            ("Analyze foot traffic and demographics", "Owner", "Week 4"),
            ("Negotiate lease terms", "Owner/Attorney", "Week 5"),
            ("Review lease with attorney", "Attorney", "Week 5"),
            ("Sign lease agreement", "Owner", "Week 6"),
            ("Get keys and access", "Owner", "Week 6"),
        ]),
        ("PHASE 3: LICENSES & PERMITS (Weeks 5-8)", [
            ("Apply for EIN (Tax ID)", "Owner", "Week 5"),
            ("Apply for state sales tax permit", "Owner", "Week 5"),
            ("Apply for business license", "Owner", "Week 5"),
            ("Health department permit (if food)", "Owner", "Week 6"),
            ("Fire safety inspection", "Fire Dept", "Week 7"),
            ("Obtain certificate of occupancy", "City", "Week 8"),
        ]),
        ("PHASE 4: STORE DESIGN & BUILD-OUT (Weeks 6-10)", [
            ("Create store layout plan", "Owner/Designer", "Week 6"),
            ("Hire contractor (if needed)", "Owner", "Week 6"),
            ("Complete electrical work", "Electrician", "Week 7"),
            ("Install flooring", "Contractor", "Week 7-8"),
            ("Paint walls and ceilings", "Contractor", "Week 8"),
            ("Install lighting", "Electrician", "Week 8"),
            ("ORDER FIXTURES AND SHELVING", "Owner", "Week 7"),
            ("Receive and install fixtures", "Owner", "Week 9"),
            ("Install signage", "Sign Company", "Week 9"),
        ]),
        ("PHASE 5: TECHNOLOGY & SYSTEMS (Weeks 8-10)", [
            ("Select and order POS system", "Owner", "Week 8"),
            ("Set up internet service", "ISP", "Week 8"),
            ("Install security cameras", "Security Co.", "Week 9"),
            ("Set up alarm system", "Security Co.", "Week 9"),
            ("Configure POS and train on usage", "Owner", "Week 10"),
            ("Set up accounting software", "Owner", "Week 10"),
        ]),
        ("PHASE 6: INVENTORY & SUPPLIERS (Weeks 8-11)", [
            ("Research and contact suppliers", "Owner", "Week 8"),
            ("Negotiate terms with vendors", "Owner", "Week 9"),
            ("Place initial inventory orders", "Owner", "Week 9"),
            ("Receive and inspect inventory", "Owner", "Week 10-11"),
            ("Tag and price all items", "Staff", "Week 11"),
            ("Organize merchandise displays", "Owner", "Week 11"),
        ]),
        ("PHASE 7: STAFFING (Weeks 8-11)", [
            ("Determine staffing needs", "Owner", "Week 8"),
            ("Write job descriptions", "Owner", "Week 8"),
            ("Post job listings", "Owner", "Week 8-9"),
            ("Interview candidates", "Owner", "Week 9-10"),
            ("Hire employees", "Owner", "Week 10"),
            ("Set up payroll", "Owner/Accountant", "Week 10"),
            ("Train employees", "Owner", "Week 11"),
        ]),
        ("PHASE 8: MARKETING & PROMOTION (Weeks 9-12)", [
            ("Create logo and branding", "Designer", "Week 9"),
            ("Build website", "Developer/Owner", "Week 9-10"),
            ("Set up social media accounts", "Owner", "Week 10"),
            ("Create Google My Business listing", "Owner", "Week 10"),
            ("Order business cards and marketing materials", "Owner", "Week 10"),
            ("Plan grand opening event", "Owner", "Week 10"),
            ("Launch social media teasers", "Owner", "Week 11-12"),
        ]),
        ("PHASE 9: FINAL PREPARATIONS (Week 12)", [
            ("Final walk-through inspection", "Owner", "Week 12"),
            ("Test all equipment and systems", "Owner", "Week 12"),
            ("Complete visual merchandising", "Owner", "Week 12"),
            ("Stock checkout area", "Owner", "Week 12"),
            ("Prepare cash drawer", "Owner", "Week 12"),
            ("Conduct soft opening (friends/family)", "Owner", "Week 12"),
        ]),
        ("PHASE 10: GRAND OPENING (Week 13)", [
            ("Final team meeting and prep", "Owner", "Opening Day"),
            ("Set up grand opening displays", "Staff", "Opening Day"),
            ("GRAND OPENING!", "Everyone", "Week 13"),
            ("Collect customer feedback", "Staff", "Week 13"),
            ("Review first day performance", "Owner", "Week 13"),
            ("Celebrate your success!", "Everyone", "Week 13"),
        ]),
    ]
    
    row = 6
    for phase_title, tasks in phases:
        create_section_header(ws, row, phase_title, 6)
        row += 1
        create_header_row(ws, row, ["Status", "Task", "Responsible", "Target Date", "Done Date", "Notes"])
        row += 1
        
        for task, responsible, target in tasks:
            status_cell = ws.cell(row=row, column=1, value="TODO")
            status_cell.border = BORDERS['thin']
            status_cell.alignment = ALIGNMENTS['center']
            status_cell.fill = FILLS['input']
            
            task_cell = ws.cell(row=row, column=2, value=task)
            task_cell.border = BORDERS['thin']
            if "FIXTURES" in task or "SHELVING" in task:
                task_cell.font = Font(bold=True, color=COLORS['accent'])
            
            ws.cell(row=row, column=3, value=responsible).border = BORDERS['thin']
            
            target_cell = ws.cell(row=row, column=4, value=target)
            target_cell.border = BORDERS['thin']
            target_cell.fill = FILLS['input']
            target_cell.alignment = ALIGNMENTS['center']
            
            ws.cell(row=row, column=5).border = BORDERS['thin']
            ws.cell(row=row, column=6).border = BORDERS['thin']
            row += 1
        
        row += 1
    
    # Footer
    row += 1
    ws.cell(row=row, column=1, value="Need quality fixtures for your new store? Contact Goodok Shopfitting!")
    ws.cell(row=row, column=1).font = Font(size=12, bold=True, color=COLORS['secondary'])
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    ws.cell(row=row, column=1, value="www.goodokshop.com | info@goodokshop.com")
    ws.cell(row=row, column=1).font = FONTS['link']
    ws.merge_cells(f'A{row}:F{row}')
    
    ws.freeze_panes = 'A2'
    
    filepath = os.path.join(output_dir, "Retail_Store_Opening_Checklist.xlsx")
    wb.save(filepath)
    return filepath


def create_inventory_tracker(output_dir):
    """创建库存管理表 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Tracker"
    
    set_column_widths(ws, {'A': 12, 'B': 35, 'C': 15, 'D': 18, 'E': 12, 'F': 12, 
                           'G': 10, 'H': 10, 'I': 12, 'J': 12, 'K': 12, 'L': 18})
    
    create_title_row(ws, 1, "RETAIL STORE INVENTORY TRACKER", 12)
    ws.cell(row=2, column=1).value = "Version 1.0 | Goodok Shopfitting"
    ws.cell(row=2, column=1).font = Font(size=10, italic=True, color='666666')
    ws.merge_cells('A2:L2')
    
    ws.cell(row=4, column=1).value = "Instructions: Enter products below. Yellow cells are editable. Status and Margin calculate automatically."
    ws.cell(row=4, column=1).font = Font(size=10, bold=True, color=COLORS['secondary'])
    ws.merge_cells('A4:L4')
    
    row = 6
    create_section_header(ws, row, "PRODUCT INVENTORY", 12)
    row += 1
    
    headers = ["SKU", "Product Name", "Category", "Supplier", "Cost ($)", "Retail ($)", 
               "Margin", "Qty", "Reorder Pt", "Order Qty", "Status", "Last Updated"]
    create_header_row(ws, row, headers)
    row += 1
    
    sample_data = [
        ("SKU001", "Sample Product A", "Category 1", "Supplier A", 10.00, 24.99, 25, 10, 50),
        ("SKU002", "Sample Product B", "Category 1", "Supplier A", 15.00, 34.99, 8, 10, 30),
        ("SKU003", "Sample Product C", "Category 2", "Supplier B", 5.00, 12.99, 0, 5, 25),
    ]
    
    start_row = row
    for i, (sku, name, cat, supplier, cost, retail, qty, reorder, order_qty) in enumerate(sample_data):
        ws.cell(row=row, column=1, value=sku).border = BORDERS['thin']
        ws.cell(row=row, column=2, value=name).border = BORDERS['thin']
        
        cat_cell = ws.cell(row=row, column=3, value=cat)
        cat_cell.border = BORDERS['thin']
        cat_cell.fill = FILLS['input']
        
        supp_cell = ws.cell(row=row, column=4, value=supplier)
        supp_cell.border = BORDERS['thin']
        supp_cell.fill = FILLS['input']
        
        cost_cell = ws.cell(row=row, column=5, value=cost)
        cost_cell.border = BORDERS['thin']
        cost_cell.fill = FILLS['input']
        cost_cell.number_format = '"$"#,##0.00'
        
        retail_cell = ws.cell(row=row, column=6, value=retail)
        retail_cell.border = BORDERS['thin']
        retail_cell.fill = FILLS['input']
        retail_cell.number_format = '"$"#,##0.00'
        
        margin_cell = ws.cell(row=row, column=7)
        margin_cell.value = f"=(F{row}-E{row})/F{row}"
        margin_cell.border = BORDERS['thin']
        margin_cell.number_format = '0%'
        
        qty_cell = ws.cell(row=row, column=8, value=qty)
        qty_cell.border = BORDERS['thin']
        qty_cell.fill = FILLS['input']
        qty_cell.alignment = ALIGNMENTS['center']
        
        reorder_cell = ws.cell(row=row, column=9, value=reorder)
        reorder_cell.border = BORDERS['thin']
        reorder_cell.fill = FILLS['input']
        reorder_cell.alignment = ALIGNMENTS['center']
        
        order_cell = ws.cell(row=row, column=10, value=order_qty)
        order_cell.border = BORDERS['thin']
        order_cell.fill = FILLS['input']
        order_cell.alignment = ALIGNMENTS['center']
        
        status_cell = ws.cell(row=row, column=11)
        status_cell.value = f'=IF(H{row}=0,"OUT OF STOCK",IF(H{row}<=I{row},"LOW STOCK","IN STOCK"))'
        status_cell.border = BORDERS['thin']
        status_cell.alignment = ALIGNMENTS['center']
        
        ws.cell(row=row, column=12, value="2024-01-01").border = BORDERS['thin']
        
        row += 1
    
    # Add empty rows for user input
    for i in range(20):
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDERS['thin']
            if col in [3, 4, 5, 6, 8, 9, 10]:
                cell.fill = FILLS['input']
            if col == 7:
                cell.value = f"=IF(F{row}>0,(F{row}-E{row})/F{row},0)"
                cell.number_format = '0%'
            if col == 11:
                cell.value = f'=IF(H{row}=0,"",IF(H{row}<=I{row},"LOW STOCK","IN STOCK"))'
            if col == 5 or col == 6:
                cell.number_format = '"$"#,##0.00'
        row += 1
    
    row += 2
    
    # Summary section
    create_section_header(ws, row, "INVENTORY SUMMARY", 12)
    row += 1
    
    ws.cell(row=row, column=1, value="Total SKUs:")
    ws.cell(row=row, column=2).value = f"=COUNTA(A{start_row}:A{row-3})"
    ws.cell(row=row, column=2).font = FONTS['total']
    row += 1
    
    ws.cell(row=row, column=1, value="Total Units:")
    ws.cell(row=row, column=2).value = f"=SUM(H{start_row}:H{row-4})"
    ws.cell(row=row, column=2).font = FONTS['total']
    row += 1
    
    ws.cell(row=row, column=1, value="Total Inventory Value:")
    ws.cell(row=row, column=2).value = f"=SUMPRODUCT(E{start_row}:E{row-5},H{start_row}:H{row-5})"
    ws.cell(row=row, column=2).font = FONTS['total']
    ws.cell(row=row, column=2).number_format = '"$"#,##0'
    row += 1
    
    ws.cell(row=row, column=1, value="Low Stock Items:")
    ws.cell(row=row, column=2).value = f'=COUNTIF(K{start_row}:K{row-6},"LOW STOCK")'
    ws.cell(row=row, column=2).font = Font(bold=True, color=COLORS['warning'])
    row += 1
    
    ws.cell(row=row, column=1, value="Out of Stock Items:")
    ws.cell(row=row, column=2).value = f'=COUNTIF(K{start_row}:K{row-7},"OUT OF STOCK")'
    ws.cell(row=row, column=2).font = Font(bold=True, color=COLORS['danger'])
    
    row += 3
    ws.cell(row=row, column=1, value="Need help organizing your store? Contact Goodok Shopfitting!")
    ws.cell(row=row, column=1).font = Font(size=12, bold=True, color=COLORS['secondary'])
    ws.merge_cells(f'A{row}:L{row}')
    
    ws.freeze_panes = 'A8'
    
    filepath = os.path.join(output_dir, "Retail_Inventory_Tracker.xlsx")
    wb.save(filepath)
    return filepath


def create_shelving_calculator(output_dir):
    """创建货架计算器 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Shelving Calculator"
    
    set_column_widths(ws, {'A': 40, 'B': 15, 'C': 12, 'D': 15, 'E': 15, 'F': 30})
    
    create_title_row(ws, 1, "RETAIL SHELVING & FIXTURE CALCULATOR", 6)
    ws.cell(row=2, column=1).value = "Version 1.0 | Goodok Shopfitting | www.goodokshop.com"
    ws.cell(row=2, column=1).font = Font(size=10, italic=True, color='666666')
    ws.merge_cells('A2:F2')
    
    ws.cell(row=4, column=1).value = "Instructions: Enter your store dimensions in the YELLOW cells. The calculator will recommend fixtures."
    ws.cell(row=4, column=1).font = Font(size=10, bold=True, color=COLORS['secondary'])
    ws.merge_cells('A4:F4')
    
    row = 6
    create_section_header(ws, row, "SECTION 1: STORE DIMENSIONS", 6)
    row += 1
    create_header_row(ws, row, ["Parameter", "Your Value", "Unit", "", "", "Notes"])
    row += 1
    
    dims = [
        ("Store Length", 50, "feet", "Front to back"),
        ("Store Width", 30, "feet", "Side to side"),
        ("Total Square Footage", "=B8*B9", "sq ft", "Auto-calculated"),
        ("Selling Floor %", 75, "%", "Exclude stockroom, office"),
        ("Selling Floor Area", "=B10*B11/100", "sq ft", "Auto-calculated"),
        ("Ceiling Height", 10, "feet", "For shelving height limits"),
    ]
    
    for param, value, unit, notes in dims:
        ws.cell(row=row, column=1, value=param).border = BORDERS['thin']
        
        val_cell = ws.cell(row=row, column=2, value=value)
        val_cell.border = BORDERS['thin']
        if not str(value).startswith('='):
            val_cell.fill = FILLS['input']
        else:
            val_cell.font = FONTS['total']
        val_cell.alignment = ALIGNMENTS['center']
        
        ws.cell(row=row, column=3, value=unit).border = BORDERS['thin']
        ws.cell(row=row, column=6, value=notes).border = BORDERS['thin']
        row += 1
    
    row += 1
    
    # Fixture Recommendations
    create_section_header(ws, row, "SECTION 2: RECOMMENDED FIXTURES (Based on Store Size)", 6)
    row += 1
    
    store_size_row = 10  # Total sq ft row
    
    ws.cell(row=row, column=1, value="Based on your store size, we recommend:")
    ws.cell(row=row, column=1).font = Font(bold=True, color=COLORS['text'])
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    create_header_row(ws, row, ["Fixture Type", "Qty Recommended", "Unit Price ($)", "Total ($)", "", "Notes"])
    row += 1
    
    fixtures = [
        ("Gondola Shelving (Double-sided)", f"=ROUND(B{store_size_row}/150,0)", 350, "Main floor displays"),
        ("Wall Shelving Units", f"=ROUND((B8+B9)*2/4,0)", 250, "Perimeter walls"),
        ("Display Cases (Glass)", f"=MAX(2,ROUND(B{store_size_row}/500,0))", 400, "High-value items"),
        ("Checkout Counter", 1, 800, ""),
        ("End Caps", f"=B{row-3}*2", 120, "2 per gondola run"),
    ]
    
    start_row = row
    for fixture, qty, price, notes in fixtures:
        ws.cell(row=row, column=1, value=fixture).border = BORDERS['thin']
        
        qty_cell = ws.cell(row=row, column=2, value=qty)
        qty_cell.border = BORDERS['thin']
        qty_cell.font = FONTS['total']
        qty_cell.alignment = ALIGNMENTS['center']
        
        price_cell = ws.cell(row=row, column=3, value=price)
        price_cell.border = BORDERS['thin']
        price_cell.fill = FILLS['input']
        price_cell.number_format = '"$"#,##0'
        
        total_cell = ws.cell(row=row, column=4)
        total_cell.value = f"=B{row}*C{row}"
        total_cell.border = BORDERS['thin']
        total_cell.number_format = '"$"#,##0'
        
        ws.cell(row=row, column=6, value=notes).border = BORDERS['thin']
        row += 1
    
    # Total
    ws.cell(row=row, column=1, value="TOTAL FIXTURE INVESTMENT")
    ws.cell(row=row, column=1).font = Font(bold=True, color=COLORS['white'])
    ws.merge_cells(f'A{row}:C{row}')
    ws.cell(row=row, column=1).fill = FILLS['primary']
    
    total_cell = ws.cell(row=row, column=4)
    total_cell.value = f"=SUM(D{start_row}:D{row-1})"
    total_cell.font = Font(bold=True, color=COLORS['white'])
    total_cell.fill = FILLS['primary']
    total_cell.number_format = '"$"#,##0'
    
    row += 2
    
    # Store type presets
    create_section_header(ws, row, "SECTION 3: STORE TYPE PRESETS (Reference)", 6)
    row += 1
    create_header_row(ws, row, ["Store Type", "Sq Ft", "Gondolas", "Wall Units", "Display Cases", "Est. Cost"])
    row += 1
    
    presets = [
        ("Convenience Store", 800, 6, 8, 2, 8500),
        ("Small Boutique", 1000, 4, 10, 4, 9200),
        ("General Retail", 1500, 10, 12, 4, 14500),
        ("Pharmacy/Health", 2000, 14, 16, 6, 21000),
        ("Large Retail", 3000, 20, 20, 8, 32000),
    ]
    
    for store_type, sqft, gondolas, wall, display, cost in presets:
        ws.cell(row=row, column=1, value=store_type).border = BORDERS['thin']
        ws.cell(row=row, column=2, value=sqft).border = BORDERS['thin']
        ws.cell(row=row, column=3, value=gondolas).border = BORDERS['thin']
        ws.cell(row=row, column=4, value=wall).border = BORDERS['thin']
        ws.cell(row=row, column=5, value=display).border = BORDERS['thin']
        cost_cell = ws.cell(row=row, column=6, value=cost)
        cost_cell.border = BORDERS['thin']
        cost_cell.number_format = '"$"#,##0'
        row += 1
    
    row += 2
    
    # Pricing guide
    create_section_header(ws, row, "SECTION 4: FIXTURE PRICING GUIDE", 6)
    row += 1
    create_header_row(ws, row, ["Fixture Type", "New Price", "Used Price", "Savings", "", "Notes"])
    row += 1
    
    pricing = [
        ("Gondola (Double-sided Starter)", 350, 175, "50%", "4ft wide, 72\" tall"),
        ("Gondola (Double-sided Add-on)", 280, 140, "50%", "Extends runs"),
        ("Wall Shelving Unit", 250, 125, "50%", "48\"x84\""),
        ("Glass Display Case", 400, 200, "50%", "Counter height"),
        ("Checkout Counter (L-shaped)", 1200, 600, "50%", "More workspace"),
        ("End Cap", 120, 60, "50%", "For promotions"),
    ]
    
    for fixture, new_price, used_price, savings, notes in pricing:
        ws.cell(row=row, column=1, value=fixture).border = BORDERS['thin']
        
        new_cell = ws.cell(row=row, column=2, value=new_price)
        new_cell.border = BORDERS['thin']
        new_cell.number_format = '"$"#,##0'
        
        used_cell = ws.cell(row=row, column=3, value=used_price)
        used_cell.border = BORDERS['thin']
        used_cell.number_format = '"$"#,##0'
        
        ws.cell(row=row, column=4, value=savings).border = BORDERS['thin']
        ws.cell(row=row, column=6, value=notes).border = BORDERS['thin']
        row += 1
    
    row += 2
    ws.cell(row=row, column=1, value="GET A FREE CUSTOM QUOTE FOR YOUR STORE!")
    ws.cell(row=row, column=1).font = Font(size=14, bold=True, color=COLORS['accent'])
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    ws.cell(row=row, column=1, value="Goodok Shopfitting - Quality Fixtures, Fair Prices")
    ws.cell(row=row, column=1).font = Font(size=12, bold=True, color=COLORS['secondary'])
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    ws.cell(row=row, column=1, value="www.goodokshop.com | info@goodokshop.com")
    ws.cell(row=row, column=1).font = FONTS['link']
    ws.merge_cells(f'A{row}:F{row}')
    
    ws.freeze_panes = 'A2'
    
    filepath = os.path.join(output_dir, "Retail_Shelving_Calculator.xlsx")
    wb.save(filepath)
    return filepath


def create_financial_projections(output_dir):
    """创建财务预测表 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Projections"
    
    set_column_widths(ws, {'A': 25, 'B': 15, 'C': 15, 'D': 15, 'E': 15, 'F': 15, 
                           'G': 15, 'H': 15, 'I': 15, 'J': 15, 'K': 15, 'L': 15, 'M': 15, 'N': 15})
    
    create_title_row(ws, 1, "RETAIL STORE FINANCIAL PROJECTIONS", 14)
    ws.cell(row=2, column=1).value = "Version 1.0 | Goodok Shopfitting"
    ws.cell(row=2, column=1).font = Font(size=10, italic=True, color='666666')
    ws.merge_cells('A2:N2')
    
    row = 4
    create_section_header(ws, row, "SECTION 1: REVENUE ASSUMPTIONS", 6)
    row += 1
    create_header_row(ws, row, ["Parameter", "Value", "Unit", "", "", "Notes"])
    row += 1
    
    assumptions = [
        ("Average Transaction Value", 35, "$", "Your target average sale"),
        ("Transactions per Day (Month 1)", 15, "count", "Conservative estimate"),
        ("Monthly Growth Rate", 10, "%", "Month-over-month"),
        ("Operating Days per Month", 26, "days", "Adjust for your schedule"),
    ]
    
    for param, value, unit, notes in assumptions:
        ws.cell(row=row, column=1, value=param).border = BORDERS['thin']
        val_cell = ws.cell(row=row, column=2, value=value)
        val_cell.border = BORDERS['thin']
        val_cell.fill = FILLS['input']
        val_cell.alignment = ALIGNMENTS['center']
        ws.cell(row=row, column=3, value=unit).border = BORDERS['thin']
        ws.cell(row=row, column=6, value=notes).border = BORDERS['thin']
        row += 1
    
    # Store references
    avg_sale_row = 6
    trans_row = 7
    growth_row = 8
    days_row = 9
    
    row += 1
    
    create_section_header(ws, row, "SECTION 2: MONTHLY COST ASSUMPTIONS", 6)
    row += 1
    create_header_row(ws, row, ["Expense Category", "Monthly ($)", "", "", "", "Notes"])
    row += 1
    
    costs = [
        ("Rent", 2500, "From your lease"),
        ("Utilities", 400, "Electric, water, internet"),
        ("Payroll", 4000, "Including taxes"),
        ("Insurance", 125, "Annual / 12"),
        ("Marketing", 300, "Ongoing promotion"),
        ("Supplies & Other", 450, "Miscellaneous"),
    ]
    
    cost_start = row
    for expense, monthly, notes in costs:
        ws.cell(row=row, column=1, value=expense).border = BORDERS['thin']
        cost_cell = ws.cell(row=row, column=2, value=monthly)
        cost_cell.border = BORDERS['thin']
        cost_cell.fill = FILLS['input']
        cost_cell.number_format = '"$"#,##0'
        ws.cell(row=row, column=6, value=notes).border = BORDERS['thin']
        row += 1
    
    ws.cell(row=row, column=1, value="TOTAL FIXED COSTS").font = FONTS['total']
    total_cost_cell = ws.cell(row=row, column=2)
    total_cost_cell.value = f"=SUM(B{cost_start}:B{row-1})"
    total_cost_cell.font = FONTS['total']
    total_cost_cell.fill = FILLS['subtotal']
    total_cost_cell.number_format = '"$"#,##0'
    fixed_cost_row = row
    
    row += 2
    
    ws.cell(row=row, column=1, value="Cost of Goods Sold (COGS) %").border = BORDERS['thin']
    cogs_cell = ws.cell(row=row, column=2, value=45)
    cogs_cell.border = BORDERS['thin']
    cogs_cell.fill = FILLS['input']
    cogs_cell.alignment = ALIGNMENTS['center']
    ws.cell(row=row, column=3, value="%").border = BORDERS['thin']
    ws.cell(row=row, column=6, value="Average product cost as % of retail").border = BORDERS['thin']
    cogs_row = row
    
    row += 2
    
    # 12-month projection
    create_section_header(ws, row, "SECTION 3: 12-MONTH REVENUE & PROFIT PROJECTION", 14)
    row += 1
    
    headers = ["Metric"] + [f"Month {i}" for i in range(1, 13)] + ["TOTAL"]
    create_header_row(ws, row, headers)
    row += 1
    
    metrics = [
        ("Transactions/Day", f"=B${trans_row}", f"=ROUND(B{row}*(1+$B${growth_row}/100),0)"),
        ("Monthly Transactions", f"=B{row-1}*$B${days_row}", f"=B{row-1}*$B${days_row}"),
        ("Revenue", f"=B{row-1}*$B${avg_sale_row}", f"=B{row-1}*$B${avg_sale_row}"),
        ("COGS", f"=B{row-1}*$B${cogs_row}/100", f"=B{row-1}*$B${cogs_row}/100"),
        ("Gross Profit", f"=B{row-2}-B{row-1}", f"=B{row-2}-B{row-1}"),
        ("Fixed Costs", f"=$B${fixed_cost_row}", f"=$B${fixed_cost_row}"),
        ("Net Profit", f"=B{row-2}-B{row-1}", f"=B{row-2}-B{row-1}"),
        ("Profit Margin", f"=IF(B{row-5}>0,B{row-1}/B{row-5},0)", f"=IF(B{row-5}>0,B{row-1}/B{row-5},0)"),
    ]
    
    metric_start = row
    for metric, m1_formula, other_formula in metrics:
        ws.cell(row=row, column=1, value=metric).border = BORDERS['thin']
        
        for col in range(2, 15):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDERS['thin']
            
            if col == 2:
                if metric == "Transactions/Day":
                    cell.value = f"=$B${trans_row}"
                elif metric == "Monthly Transactions":
                    cell.value = f"=B{row-1}*$B${days_row}"
                elif metric == "Revenue":
                    cell.value = f"=B{row-1}*$B${avg_sale_row}"
                elif metric == "COGS":
                    cell.value = f"=B{row-1}*$B${cogs_row}/100"
                elif metric == "Gross Profit":
                    cell.value = f"=B{row-2}-B{row-1}"
                elif metric == "Fixed Costs":
                    cell.value = f"=$B${fixed_cost_row}"
                elif metric == "Net Profit":
                    cell.value = f"=B{row-2}-B{row-1}"
                elif metric == "Profit Margin":
                    cell.value = f"=IF(B{row-5}>0,B{row-1}/B{row-5},0)"
            elif col == 14:  # Total column
                if metric in ["Profit Margin"]:
                    cell.value = f"=AVERAGE(B{row}:M{row})"
                else:
                    cell.value = f"=SUM(B{row}:M{row})"
            else:
                prev_col = chr(ord('A') + col - 2)
                curr_col = chr(ord('A') + col - 1)
                if metric == "Transactions/Day":
                    cell.value = f"=ROUND({prev_col}{row}*(1+$B${growth_row}/100),0)"
                elif metric == "Monthly Transactions":
                    cell.value = f"={curr_col}{row-1}*$B${days_row}"
                elif metric == "Revenue":
                    cell.value = f"={curr_col}{row-1}*$B${avg_sale_row}"
                elif metric == "COGS":
                    cell.value = f"={curr_col}{row-1}*$B${cogs_row}/100"
                elif metric == "Gross Profit":
                    cell.value = f"={curr_col}{row-2}-{curr_col}{row-1}"
                elif metric == "Fixed Costs":
                    cell.value = f"=$B${fixed_cost_row}"
                elif metric == "Net Profit":
                    cell.value = f"={curr_col}{row-2}-{curr_col}{row-1}"
                elif metric == "Profit Margin":
                    cell.value = f"=IF({curr_col}{row-5}>0,{curr_col}{row-1}/{curr_col}{row-5},0)"
            
            if metric in ["Revenue", "COGS", "Gross Profit", "Fixed Costs", "Net Profit"]:
                cell.number_format = '"$"#,##0'
            elif metric == "Profit Margin":
                cell.number_format = '0.0%'
        
        if metric == "Net Profit":
            ws.cell(row=row, column=1).font = FONTS['total']
            for col in range(2, 15):
                ws.cell(row=row, column=col).font = FONTS['total']
        
        row += 1
    
    row += 2
    
    # Break-even analysis
    create_section_header(ws, row, "SECTION 4: BREAK-EVEN ANALYSIS", 6)
    row += 1
    
    ws.cell(row=row, column=1, value="Total Fixed Costs (Monthly):").border = BORDERS['thin']
    ws.cell(row=row, column=2).value = f"=$B${fixed_cost_row}"
    ws.cell(row=row, column=2).number_format = '"$"#,##0'
    row += 1
    
    ws.cell(row=row, column=1, value="Gross Margin:").border = BORDERS['thin']
    ws.cell(row=row, column=2).value = f"=1-$B${cogs_row}/100"
    ws.cell(row=row, column=2).number_format = '0%'
    row += 1
    
    ws.cell(row=row, column=1, value="Break-Even Revenue (Monthly):").border = BORDERS['thin']
    ws.cell(row=row, column=2).value = f"=$B${fixed_cost_row}/(1-$B${cogs_row}/100)"
    ws.cell(row=row, column=2).number_format = '"$"#,##0'
    ws.cell(row=row, column=2).font = FONTS['total']
    row += 1
    
    ws.cell(row=row, column=1, value="Break-Even Transactions/Day:").border = BORDERS['thin']
    ws.cell(row=row, column=2).value = f"=B{row-1}/($B${avg_sale_row}*$B${days_row})"
    ws.cell(row=row, column=2).number_format = '0.0'
    ws.cell(row=row, column=2).font = FONTS['total']
    
    row += 3
    ws.cell(row=row, column=1, value="Save on fixtures to improve your margins!")
    ws.cell(row=row, column=1).font = Font(size=12, bold=True, color=COLORS['secondary'])
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    ws.cell(row=row, column=1, value="www.goodokshop.com | info@goodokshop.com")
    ws.cell(row=row, column=1).font = FONTS['link']
    ws.merge_cells(f'A{row}:F{row}')
    
    ws.freeze_panes = 'B5'
    
    filepath = os.path.join(output_dir, "Retail_Financial_Projections.xlsx")
    wb.save(filepath)
    return filepath


if __name__ == "__main__":
    output_dir = r"E:\retail-miniapp-backend\goodok-website\digital-products"
    
    print("Generating remaining Excel templates...")
    
    print("  Creating Store Opening Checklist...")
    filepath = create_store_opening_checklist(output_dir)
    print(f"  Saved: {filepath}")
    
    print("  Creating Inventory Tracker...")
    filepath = create_inventory_tracker(output_dir)
    print(f"  Saved: {filepath}")
    
    print("  Creating Shelving Calculator...")
    filepath = create_shelving_calculator(output_dir)
    print(f"  Saved: {filepath}")
    
    print("  Creating Financial Projections...")
    filepath = create_financial_projections(output_dir)
    print(f"  Saved: {filepath}")
    
    print("\nAll Excel templates generated successfully!")
