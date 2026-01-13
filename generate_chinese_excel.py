"""
零售店规划 Excel 套装中文版生成器
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os

# 样式定义
COLORS = {
    'primary': '1E3A5F', 'secondary': '2E86AB', 'accent': 'F5A623',
    'success': '27AE60', 'warning': 'F39C12', 'danger': 'E74C3C',
    'light': 'F8F9FA', 'white': 'FFFFFF', 'text': '2C3E50', 'input_bg': 'FFFDE7',
}

def get_fonts():
    return {
        'title': Font(name='微软雅黑', size=18, bold=True, color=COLORS['white']),
        'section': Font(name='微软雅黑', size=14, bold=True, color=COLORS['white']),
        'header': Font(name='微软雅黑', size=11, bold=True, color=COLORS['text']),
        'normal': Font(name='微软雅黑', size=10, color=COLORS['text']),
        'total': Font(name='微软雅黑', size=11, bold=True, color=COLORS['primary']),
        'link': Font(name='微软雅黑', size=10, color=COLORS['secondary'], underline='single'),
    }

FILLS = {
    'primary': PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid'),
    'secondary': PatternFill(start_color=COLORS['secondary'], end_color=COLORS['secondary'], fill_type='solid'),
    'header': PatternFill(start_color=COLORS['light'], end_color=COLORS['light'], fill_type='solid'),
    'input': PatternFill(start_color=COLORS['input_bg'], end_color=COLORS['input_bg'], fill_type='solid'),
    'subtotal': PatternFill(start_color='E8F4FD', end_color='E8F4FD', fill_type='solid'),
}

BORDERS = {
    'thin': Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    ),
}

ALIGNMENTS = {
    'center': Alignment(horizontal='center', vertical='center'),
    'left': Alignment(horizontal='left', vertical='center'),
}

def setup_sheet(ws, title, subtitle, col_widths, title_cols):
    fonts = get_fonts()
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=title_cols)
    cell = ws.cell(row=1, column=1)
    cell.value = title
    cell.font = fonts['title']
    cell.fill = FILLS['primary']
    cell.alignment = ALIGNMENTS['center']
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=title_cols)
    cell = ws.cell(row=2, column=1)
    cell.value = subtitle
    cell.font = Font(name='微软雅黑', size=10, italic=True, color='666666')
    ws.row_dimensions[2].height = 20

def create_section(ws, row, title, headers, cols):
    fonts = get_fonts()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = title
    cell.font = fonts['section']
    cell.fill = FILLS['secondary']
    cell.alignment = ALIGNMENTS['center']
    ws.row_dimensions[row].height = 28
    row += 1
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = fonts['header']
        cell.fill = FILLS['header']
        cell.border = BORDERS['thin']
        cell.alignment = ALIGNMENTS['center']
    ws.row_dimensions[row].height = 22
    return row + 1

# --- 1. 成本计算器 ---
def make_cost_calc(output_dir):
    wb = Workbook()
    ws = wb.active
    ws.title = "开店成本计算器"
    setup_sheet(ws, "🏪 零售店开店成本计算器", "版本 1.0 | Goodok 货架策划 | www.goodokshop.com", {'A': 30, 'B': 40, 'C': 18, 'D': 45}, 4)
    
    row = 4
    ws.cell(row=row, column=1, value="📝 说明：在黄色单元格填入估算金额。公式会自动汇总。").font = Font(name='微软雅黑', size=10, bold=True, color=COLORS['secondary'])
    
    sections = [
        ("📍 第一部分：选址与租约费用", ["费用类别", "具体项目", "预计成本 (¥)", "备注"], [
            ("房租", "租赁押金 (通常2-3个月)", 30000, ""),
            ("房租", "首月房租", 15000, ""),
            ("中介", "中介佣金", 7500, ""),
            ("法务", "合同审阅/律师费", 2000, ""),
        ]),
        ("🔨 第二部分：装修与改造费用", ["费用类别", "具体项目", "预计成本 (¥)", "备注"], [
            ("装修", "地面铺设", 15000, ""),
            ("装修", "墙面喷涂/软装", 8000, ""),
            ("灯光", "灯具采购及安装", 12000, ""),
            ("电路", "电路改造", 10000, ""),
            ("招牌", "外部招牌/发光字", 10000, ""),
        ]),
    ]
    
    row = 6
    subtotals = []
    for sec_title, headers, data in sections:
        row = create_section(ws, row, sec_title, headers, 4)
        start_row = row
        for cat, item, cost, notes in data:
            ws.cell(row=row, column=1, value=cat).border = BORDERS['thin']
            ws.cell(row=row, column=2, value=item).border = BORDERS['thin']
            c = ws.cell(row=row, column=3, value=cost)
            c.fill = FILLS['input']; c.border = BORDERS['thin']; c.number_format = '"¥"#,##0'
            ws.cell(row=row, column=4, value=notes).border = BORDERS['thin']
            row += 1
        
        ws.cell(row=row, column=1, value="本项小计").font = get_fonts()['total']
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        total_c = ws.cell(row=row, column=3, value=f"=SUM(C{start_row}:C{row-1})")
        total_c.font = get_fonts()['total']; total_c.fill = FILLS['subtotal']; total_c.border = BORDERS['thin']; total_c.number_format = '"¥"#,##0'
        subtotals.append(f"C{row}")
        row += 2

    # 货架部分特殊处理
    row = create_section(ws, row, "🛒 第三部分：货架与设备 ⭐", ["项目", "数量", "单价 (¥)", "合计 (¥)"], 4)
    start_row = row
    fixtures = [("双面中岛货架 (1.2米宽)", 10, 2200), ("单面靠墙货架", 8, 1500), ("玻璃展示柜", 4, 2500), ("收银台", 1, 4500)]
    for item, qty, price in fixtures:
        ws.cell(row=row, column=1, value=item).border = BORDERS['thin']
        q = ws.cell(row=row, column=2, value=qty); q.fill = FILLS['input']; q.border = BORDERS['thin']; q.alignment = ALIGNMENTS['center']
        p = ws.cell(row=row, column=3, value=price); p.fill = FILLS['input']; p.border = BORDERS['thin']; p.number_format = '"¥"#,##0'
        t = ws.cell(row=row, column=4, value=f"=B{row}*C{row}"); t.border = BORDERS['thin']; t.number_format = '"¥"#,##0'
        row += 1
    
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value="本项小计").font = get_fonts()['total']
    total_c = ws.cell(row=row, column=4, value=f"=SUM(D{start_row}:D{row-1})")
    total_c.font = get_fonts()['total']; total_c.fill = FILLS['subtotal']; total_c.border = BORDERS['thin']; total_c.number_format = '"¥"#,##0'
    subtotals.append(f"D{row}")
    row += 3

    create_section(ws, row, "🎯 开店总成本汇总", ["项目", "", "", "金额 (¥)"], 4)
    row += 1
    summary = [("📍 空间与租约", subtotals[0]), ("🔨 装修与改造", subtotals[1]), ("🛒 货架与设备", subtotals[2])]
    for label, ref in summary:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.cell(row=row, column=1, value=label).border = BORDERS['thin']
        v = ws.cell(row=row, column=4, value=f"={ref}"); v.border = BORDERS['thin']; v.number_format = '"¥"#,##0'
        row += 1
    
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value="✨ 预估开店总投资").font = Font(name='微软雅黑', size=14, bold=True, color=COLORS['white'])
    ws.cell(row=row, column=1).fill = FILLS['primary']
    grand_total = ws.cell(row=row, column=4, value=f"=SUM(D{row-3}:D{row-1})")
    grand_total.font = Font(name='微软雅黑', size=14, bold=True, color=COLORS['white']); grand_total.fill = FILLS['primary']; grand_total.number_format = '"¥"#,##0'
    ws.row_dimensions[row].height = 30

    wb.save(os.path.join(output_dir, "零售店开店成本计算器_中文.xlsx"))

# --- 2. 任务清单 ---
def make_checklist(output_dir):
    wb = Workbook()
    ws = wb.active
    ws.title = "开店任务清单"
    setup_sheet(ws, "✅ 零售店开店筹备清单", "版本 1.0 | Goodok 货架策划", {'A': 12, 'B': 50, 'C': 15, 'D': 15, 'E': 35}, 5)
    
    phases = [
        ("阶段 1：调研与规划 (第1-4周)", [("明确目标市场和定位", "负责人"), ("竞品调研与选址分析", "负责人"), ("编写商业计划书", "负责人")]),
        ("阶段 2：店铺选址与租赁 (第4-6周)", [("寻找合适商铺", "负责人"), ("签订租赁合同", "负责人"), ("办理营业执照", "负责人")]),
        ("阶段 3：设计与装修 (第6-10周)", [("空间布局规划", "货架专家"), ("订购货架与展示柜", "Goodok"), ("完成内部装修", "施工方")]),
    ]
    
    row = 6
    for title, tasks in phases:
        row = create_section(ws, row, title, ["状态", "任务描述", "负责人", "目标日期", "备注"], 5)
        for t_desc, resp in tasks:
            ws.cell(row=row, column=1, value="待办").alignment = ALIGNMENTS['center']
            ws.cell(row=row, column=2, value=t_desc)
            ws.cell(row=row, column=3, value=resp)
            for c in range(1, 6): ws.cell(row=row, column=c).border = BORDERS['thin']
            row += 1
        row += 1
    wb.save(os.path.join(output_dir, "零售店开店筹备清单_中文.xlsx"))

# --- 3. 库存管理 ---
def make_inventory(output_dir):
    wb = Workbook()
    ws = wb.active
    ws.title = "库存管理表"
    setup_sheet(ws, "📦 零售店内部库存管理表", "版本 1.0 | Goodok 货架策划", {'A': 15, 'B': 30, 'C': 15, 'D': 12, 'E': 12, 'F': 12, 'G': 15}, 7)
    
    row = create_section(ws, row=6, title="商品库存清单", headers=["商品编号", "商品名称", "分类", "进货价(¥)", "销售价(¥)", "当前库存", "库存状态"], cols=7)
    for i in range(10):
        ws.cell(row=row, column=1, value=f"GDK-{100+i}")
        ws.cell(row=row, column=4).fill = FILLS['input']; ws.cell(row=row, column=4).number_format = '"¥"#,##0'
        ws.cell(row=row, column=5).fill = FILLS['input']; ws.cell(row=row, column=5).number_format = '"¥"#,##0'
        ws.cell(row=row, column=6, value=20).fill = FILLS['input']
        ws.cell(row=row, column=7, value='=IF(F%d<5,"需补货","充足")' % row)
        for c in range(1, 8): ws.cell(row=row, column=c).border = BORDERS['thin']
        row += 1
    wb.save(os.path.join(output_dir, "零售店库存管理表_中文.xlsx"))

# --- 4. 财务预测 ---
def make_finance(output_dir):
    wb = Workbook()
    ws = wb.active
    setup_sheet(ws, "💰 零售店12个月财务预测表", "版本 1.0 | Goodok 货架策划", {'A': 20, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 12}, 6)
    row = create_section(ws, 6, "核心盈利预测 (每月)", ["月份", "预计收入", "成本 (COGS)", "固定支出", "净利润", "利润率"], 6)
    for i in range(1, 13):
        ws.cell(row=row, column=1, value=f"{i}月")
        ws.cell(row=row, column=2, value=50000).fill = FILLS['input']; ws.cell(row=row, column=2).number_format = '"¥"#,##0'
        ws.cell(row=row, column=3, value=f"=B{row}*0.5").number_format = '"¥"#,##0'
        ws.cell(row=row, column=4, value=15000).fill = FILLS['input']; ws.cell(row=row, column=4).number_format = '"¥"#,##0'
        ws.cell(row=row, column=5, value=f"=B{row}-C{row}-D{row}").number_format = '"¥"#,##0'
        ws.cell(row=row, column=6, value=f"=E{row}/B{row}").number_format = '0%'
        for c in range(1, 7): ws.cell(row=row, column=c).border = BORDERS['thin']
        row += 1
    wb.save(os.path.join(output_dir, "零售店财务预测表_中文.xlsx"))

# --- 5. 货架计算器 ---
def make_shelving(output_dir):
    wb = Workbook()
    ws = wb.active
    setup_sheet(ws, "📐 零售店货架及布局计算器", "专业货架规划 - Goodok Shopfitting", {'A': 30, 'B': 20, 'C': 40}, 3)
    row = create_section(ws, 6, "店铺空间参数", ["参数名称", "输入值", "备注"], 3)
    params = [("店铺长度 (米)", 15, "墙到墙距离"), ("店铺宽度 (米)", 10, ""), ("总面积 (平米)", "=B8*B9", "自动计算")]
    for p, v, n in params:
        ws.cell(row=row, column=1, value=p)
        c = ws.cell(row=row, column=2, value=v); c.border = BORDERS['thin']
        if not str(v).startswith('='): c.fill = FILLS['input']
        ws.cell(row=row, column=3, value=n)
        row += 1
    wb.save(os.path.join(output_dir, "零售店货架布局计算器_中文.xlsx"))

if __name__ == "__main__":
    out = r"E:\retail-miniapp-backend\goodok-website\digital-products"
    if not os.path.exists(out): os.makedirs(out)
    make_cost_calc(out)
    make_checklist(out)
    make_inventory(out)
    make_finance(out)
    make_shelving(out)
    print("中文版 Excel 模板生成成功！")
